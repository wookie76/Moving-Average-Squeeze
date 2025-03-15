import numpy as np
import yfinance as yf
from rich.console import Console
from rich.table import Table
from rich.text import Text
from tqdm import tqdm
from pydantic import BaseModel, Field, field_validator
from typing import List, Literal
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font


class HullMovingAverage:
    """Calculates the Hull Moving Average (HMA) - Optimized."""

    def __init__(self, prices: np.ndarray, n: int):
        # No need to store prices, calculate on the fly.  Data should be float64 for precision.
        self.n = n
        if not isinstance(prices, np.ndarray):
            raise TypeError("prices must be a NumPy array")
        if prices.ndim != 1:
            raise ValueError("prices must be a 1-dimensional array")
        if not np.issubdtype(prices.dtype, np.floating):
            self.prices = prices.astype(np.float64)
        else:
            self.prices = prices


    def wma(self, periods: int) -> np.ndarray:
        """Calculates the Weighted Moving Average (WMA) - Optimized."""

        if len(self.prices) < periods:
            return np.array([np.nan] * len(self.prices))

        weights = np.arange(1, periods + 1)
        weights_sum = periods * (periods + 1) / 2  # Sum of arithmetic series
        # Ensure that self.prices is flattened before convolution
        wma = np.convolve(self.prices.flatten(), weights / weights_sum, mode="valid")
        # Pad with NaNs *at the beginning*, as convolution removes 'periods - 1' elements.
        return np.pad(wma, (periods - 1, 0), constant_values=np.nan)


    def calculate(self) -> np.ndarray:
        """Calculates the Hull Moving Average - Optimized."""
        n = self.n
        if len(self.prices) < n:
            return np.array([np.nan] * len(self.prices))

        n_half = n // 2
        n_sqrt = int(np.floor(np.sqrt(n)))

        wma_half = self.wma(n_half)
        wma_full = self.wma(n)

        # Calculate raw_hma only where both wma_half and wma_full are not NaN
        raw_hma = np.where(
            ~np.isnan(wma_half) & ~np.isnan(wma_full),
            (2 * wma_half) - wma_full,
            np.nan
        )

        # Create a new instance with valid raw_hma
        hma_calculator = HullMovingAverage(raw_hma[~np.isnan(raw_hma)], n_sqrt)
        result = hma_calculator.wma(n_sqrt)
        # Pad the result with NaNs to match the original length.
        padding_length = len(self.prices) - len(result)
        return np.pad(result, (padding_length, 0), constant_values=np.nan)



def get_closing_prices(ticker: str, period: str,
                       interval: str) -> np.ndarray | None:
    """Retrieves historical closing prices using yfinance - Optimized."""
    try:
        data = yf.download(ticker,
                           period=period,
                           interval=interval,
                           auto_adjust=True,
                           progress=False)  # Disable internal tqdm
        # Directly use a NumPy array for efficiency and ensure float64
        closing_prices = data["Close"].to_numpy(dtype=np.float64)
        if closing_prices.size == 0:  # Check for empty data
            print(f"No data found for {ticker}")
            return None
        return closing_prices.flatten()  #  Flatten the array here
    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")
        return None


def calculate_hma_squeeze_signals(hma_21: np.ndarray, hma_89: np.ndarray,
                                  squeeze_window: int) -> tuple[np.ndarray,
                                                               np.ndarray,
                                                               np.ndarray]:
    """Calculates squeeze and entry/exit signals - Optimized."""
    # Input validation
    if not (isinstance(hma_21, np.ndarray) and isinstance(hma_89, np.ndarray)):
        raise TypeError("hma_21 and hma_89 must be NumPy arrays")
    if hma_21.ndim != 1 or hma_89.ndim != 1:
        raise ValueError("hma_21 and hma_89 must be 1-dimensional arrays")
    if len(hma_21) != len(hma_89):
        raise ValueError("hma_21 and hma_89 must have the same length")

    # Calculate the absolute difference *first*.  This is crucial.
    hma_diff = np.abs(hma_21 - hma_89)

    # *Now* replace any NaN values in hma_diff with infinity.
    hma_diff = np.nan_to_num(hma_diff, nan=np.inf)


    # More efficient threshold calculation (avoid full convolution)
    valid_len = len(hma_diff) - squeeze_window + 1
    squeeze_threshold = np.full(len(hma_diff), np.inf)  # Initialize with infinity

    if valid_len > 0:
        windowed_sum = np.convolve(hma_diff, np.ones(squeeze_window), mode="valid")
        squeeze_threshold[-valid_len:] = windowed_sum / squeeze_window

    squeeze = hma_diff <= squeeze_threshold

    # Vectorized entry and exit signal calculation.  Handle NaNs properly.
    hma_21_valid = np.nan_to_num(hma_21, nan=-np.inf)  # For comparisons
    hma_89_valid = np.nan_to_num(hma_89, nan=-np.inf)

    #  Use .flatten() on the arrays before np.roll
    entry = (hma_21_valid > hma_89_valid) & (np.roll(hma_21_valid.flatten(), 1, axis=0) <= np.roll(hma_89_valid.flatten(), 1, axis=0)) & squeeze
    exit_signal = (hma_21_valid < hma_89_valid) & (np.roll(hma_21_valid.flatten(), 1, axis=0) >= np.roll(hma_89_valid.flatten(), 1, axis=0))

    entry[0] = False  # Avoid first-element issues
    exit_signal[0] = False

    return squeeze, entry, exit_signal



def backtest_strategy(
    prices: np.ndarray,
    entry_signals: np.ndarray,
    exit_signals: np.ndarray,
    initial_capital: float = 10000.0,
) -> tuple[float, List[float]]:
    """Backtests the HMA squeeze strategy - Optimized."""
    # Input validation
    if not isinstance(prices, np.ndarray):
        raise TypeError("prices must be a NumPy array")
    if prices.ndim != 1:
        raise ValueError("prices must be a 1-dimensional array")
    if not (isinstance(entry_signals, np.ndarray) and isinstance(exit_signals, np.ndarray)):
        raise TypeError("entry_signals and exit_signals must be NumPy arrays")
    if entry_signals.ndim != 1 or exit_signals.ndim != 1:
         raise ValueError("entry_signals and exit_signals must be 1-dimensional arrays")

    position = 0  # 0: no position, 1: long
    cash = initial_capital
    portfolio_value = []
    shares = 0

    # Vectorized backtesting loop
    for i in range(len(prices)):
        if entry_signals[i] and position == 0:
            shares = cash / prices[i]
            cash = 0
            position = 1
        elif exit_signals[i] and position == 1:
            cash = shares * prices[i]
            shares = 0
            position = 0
        portfolio_value.append(cash + shares * prices[i])

    final_value = portfolio_value[-1] if portfolio_value else initial_capital # Handle empty case.
    return final_value, portfolio_value


def calculate_cagr(begin_value: float, end_value: float,
                   num_periods: float) -> float | None:
    """Calculates the Compound Annual Growth Rate (CAGR)."""
    if begin_value <= 0 or num_periods <= 0:
        return None
    # No need to check for None, result is always float or error
    return (end_value / begin_value)**(1 / num_periods) - 1



class Config(BaseModel):
    tickers_file: str = Field("all_tickers.txt",
                              description="Path to the file with tickers")
    period: str = Field("2y", description="Data period")
    interval: str = Field("1h", description="Data interval")
    squeeze_window: int = Field(21, description="Squeeze window", ge=5)  # Default to 21
    output_file: str = Field("hma_squeeze_results.xlsx", description="Output Excel file name")

    @field_validator("period")
    def validate_period(cls, v):
        if not any(char.isdigit() for char in v):
            raise ValueError("Period must contain a number")
        if not v.endswith(("d", "wk", "mo", "y")):
            raise ValueError("Invalid period format")
        return v

    @field_validator("interval")
    def validate_interval(cls, v):
        if not any(char.isdigit() for char in v):
            raise ValueError("Interval must contain a number")
        if not v.endswith(("m", "h", "d", "wk", "mo")):
            raise ValueError("Invalid interval format")
        return v

    def load_tickers(self) -> List[str]:
        """Loads tickers from the specified file."""
        tickers_path = Path(self.tickers_file)
        if not tickers_path.exists():
            raise FileNotFoundError(f"Tickers file not found: {self.tickers_file}")
        with open(tickers_path, "r") as f:
            return [line.strip().upper() for line in f]


def main():
    """Calculates HMAs, performs backtesting, and presents results, exporting to Excel."""

    config = Config()
    tickers = config.load_tickers()
    console = Console()

    results_data = []  # Store results for Excel export

    for ticker in tqdm(tickers, desc="Processing Tickers"):
        closing_prices = get_closing_prices(ticker, config.period,
                                            config.interval)
        if closing_prices is None:
            console.print(
                f"[bold red]Error: Could not fetch data for {ticker}[/]")
            continue

        # Create HMA instances with the closing prices.
        hma_21_calculator = HullMovingAverage(closing_prices, 21)
        hma_89_calculator = HullMovingAverage(closing_prices, 89)
        hma_21 = hma_21_calculator.calculate()
        hma_89 = hma_89_calculator.calculate()


        squeeze, entry, exit_signal = calculate_hma_squeeze_signals(
            hma_21, hma_89, config.squeeze_window)

        final_value_strategy, _ = backtest_strategy(closing_prices, entry,
                                                    exit_signal)

        # Handle cases where there are no valid closing prices.
        if closing_prices.size > 0:
          initial_price = closing_prices[0]
          final_price_bh = closing_prices[-1]
          final_value_bh = 10000.0 * (final_price_bh / initial_price)
        else:
            initial_price = final_price_bh = final_value_bh = 0.0  # Set to 0 if no data

        num_years = (
            float(config.period.replace("y", ""))
            if "y" in config.period else
            (float(config.period.replace("mo", "")) / 12)
            if "mo" in config.period else
            (float(config.period.replace("wk", "")) / 52)
            if "wk" in config.period else
            (float(config.period.replace("d", "")) / 365)
            if "d" in config.period else 1 # Should never be < 1
        )

        cagr_strategy = calculate_cagr(10000.0, final_value_strategy, num_years)
        cagr_bh = calculate_cagr(10000.0, final_value_bh, num_years)

        # Determine the last signal, handling edge cases using .sum() for conciseness
        last_entry = entry.sum()
        last_exit = exit_signal.sum()

        if last_entry > last_exit:
             last_signal = "Entry"
        elif last_exit > last_entry:
            last_signal = "Exit"
        else:
            last_signal = "None"

        # Append results to the data list
        results_data.append({
            "Ticker": ticker,
            "Final Value (Strategy)": f"${final_value_strategy:.2f}",
            "CAGR (Strategy)": f"{cagr_strategy:.2%}" if cagr_strategy is not None else "N/A",
            "Final Value (B&H)": f"${final_value_bh:.2f}",
            "CAGR (B&H)": f"{cagr_bh:.2%}" if cagr_bh is not None else "N/A",
            "Last Signal": last_signal
        })

    # Create a new Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "HMA Squeeze Results"

    # Write the header row
    header = list(results_data[0].keys())
    for col_num, header_text in enumerate(header, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write the data rows
    for row_num, row_data in enumerate(results_data, 2):
        for col_num, key in enumerate(header, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=row_data[key])
            if col_num in (2,4): #numeric columns
               cell.alignment = Alignment(horizontal='right')
            if col_num in (3,5):
                cell.alignment = Alignment(horizontal='right')
            if col_num == 6:
                cell.alignment = Alignment(horizontal='center')


    # Adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)  # Get column letter
        for cell in column:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    workbook.save(config.output_file)
    console.print(f"[bold green]Results exported to {config.output_file}[/]")



if __name__ == "__main__":
    main()